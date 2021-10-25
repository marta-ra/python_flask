from flask import Flask, render_template, request
import telnetlib
from netmiko import ConnectHandler
import csv
import re
import time
import os
from openpyxl import load_workbook
from log_pass import LOGIN, PASSWORD

app = Flask(__name__)

LOGIN = LOGIN
PASSWORD = PASSWORD


def to_bytes(line):
    return f"{line}".encode("utf-8")

def send_command(ip_device, login, command):
    connect = telnetlib.Telnet(to_bytes(ip_device))
    connect.write(to_bytes(command) + b'\n')
    output = connect.read_until(b']')
    writer_log_file(ip_device, login, command, output)

def writer_log_file(host, login, command, output):
    name_file = host + '_' + login + '_' + 'logfile.csv'
    try:
        with open(name_file, 'a', newline='') as csvfile:
            spamwriter = csv.writer(csvfile, delimiter=';')
            spamwriter.writerow([host, login, str(command), str(output)])
            return None
    except:
        return 'File error'


def check_login_ip_on_router(ip_router, vlan):

    device = {
        "device_type": "cisco_ios",
        "ip": ip_router,
        "username": LOGIN,
        "password": PASSWORD,
        }
    print(f'connet to {ip_router}')
    try:
        with ConnectHandler(**device) as ssh:
            ssh.enable()

            output_login_vlan = ssh.send_command('show running-config-interface vlan' + vlan)
            print(output_login_vlan)
            check_login_ip = re.findall(r"ip address\s\b\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\b\s\b\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\b", output_login_vlan) 
            print(check_login_ip)
            for vaule in check_login_ip:
                login_ip_on_router = vaule
            ConnectHandler(**device).disconnect()
            return login_ip_on_router
    except:
        print('Не удалось подключиться к роутеру')
        login_ip_on_aggregation = 'Не удалось подключиться к роутеру'
        return login_ip_on_aggregation


def search_login_inf(lookFor_client):
    # Поиск по файлу 
    path = r'path'
    tg_xlsx = load_workbook(os.path.join(path, "data.xlsx"), read_only=True)
    tg_sheet = tg_xlsx.active
    max_row_not_empty = tg_sheet.max_row
    # При нахождении логина в файле извлекаются ячейки с соответствующими значениями 
    check_find_row = 0
    for rownum in range(1, max_row_not_empty + 1):
        if check_find_row == 0:
            search_login = str(tg_sheet.cell(row=rownum, column=2).value)            
            if search_login == lookFor_client:     
                check_find_row = 1
                switch_model = str(tg_sheet.cell(row=rownum, column=14).value)
                ip_device= str(tg_sheet.cell(row=rownum, column=1).value)
                client_switch_name = str(tg_sheet.cell(row=rownum, column=7).value)
                client_login = search_login
        elif check_find_row == 1:
            break
    return [ip_device, client_switch_name, client_login, switch_model]



@app.route('/')
def mac_input_data():
    return render_template('index.html')



@app.route('/result_check_login', methods=['POST'])
def client_data():
    lookFor_client = request.form['login_check']
    list_login_inf = search_login_inf(lookFor_client)
    return render_template("result.html", result = list_login_inf)



if __name__ == '__main__':
    app.run(host='0.0.0.0',debug='True')
